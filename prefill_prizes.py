import json
from openpyxl import load_workbook

with open('data/weekly_highs.json') as f:
    team_highs = json.load(f)
with open('data/player_highs.json') as f:
    player_highs = json.load(f)
with open('data/blowouts_streaks.json') as f:
    blowouts = json.load(f)

# Build lookup dicts
team_weekly = {}
for year_str, data in team_highs.items():
    for week_str, info in data['weekly_highs'].items():
        team_weekly[(int(year_str), int(week_str))] = info

team_season = {}
for year_str, data in team_highs.items():
    team_season[int(year_str)] = data['season_high']

player_weekly = {}
for entry in player_highs.get('weekly_highs', []):
    if entry['points'] > 0 and not entry.get('is_playoffs'):
        key = (entry['year'], entry['week'])
        if key not in player_weekly or entry['points'] > player_weekly[key]['points']:
            player_weekly[key] = entry

player_season = {}
for entry in player_highs.get('weekly_highs', []):
    if entry['points'] > 0 and not entry.get('is_playoffs'):
        year = entry['year']
        if year not in player_season or entry['points'] > player_season[year]['points']:
            player_season[year] = entry

wb = load_workbook('GFN_Fantasy_Finances_v3.xlsx')

for year in range(2011, 2026):
    if str(year) not in wb.sheetnames:
        continue

    ws = wb[str(year)]
    filled = []
    blowout = blowouts.get(str(year), {}).get('blowout', {})
    streak = blowouts.get(str(year), {}).get('win_streak', {})

    for row in ws.iter_rows():
        for c in row:
            val = str(c.value).strip() if c.value else ""

            # Weekly High Score (team)
            if val == "Weekly High Score":
                week_raw = ws.cell(row=c.row, column=1).value
                if week_raw is not None:
                    week_num = int(float(week_raw))
                    if week_num <= 14:
                        info = team_weekly.get((year, week_num))
                        if info:
                            ws.cell(row=c.row, column=3).value = info['owner']
                            filled.append(f"Wk {week_num} Team High: {info['owner']} ({info['points']} pts)")

            # Single Player Weekly High Score
            if val == "Single Player Weekly High Score":
                week_raw = ws.cell(row=c.row, column=1).value
                if week_raw is not None:
                    week_num = int(float(week_raw))
                    info = player_weekly.get((year, week_num))
                    if info:
                        ws.cell(row=c.row, column=3).value = info['owner']
                        ws.cell(row=c.row, column=5).value = f"{info['player']} — {info['points']} pts"
                        filled.append(f"Wk {week_num} Player High: {info['player']} ({info['owner']}) {info['points']} pts")

            # Season High Points (team)
            if val == "Season High Points":
                winner_cell = ws.cell(row=c.row, column=3)
                if not winner_cell.value:
                    info = team_season.get(year)
                    if info and info.get('owner'):
                        winner_cell.value = info['owner']
                        ws.cell(row=c.row, column=5).value = f"{info['points']} pts"
                        filled.append(f"Season Team High: {info['owner']} ({info['points']} pts)")

            # Single Player Season High Score
            if val == "Single Player Season High Score":
                winner_cell = ws.cell(row=c.row, column=3)
                if not winner_cell.value:
                    info = player_season.get(year)
                    if info:
                        winner_cell.value = info['owner']
                        ws.cell(row=c.row, column=5).value = f"{info['player']} — {info['points']} pts"
                        filled.append(f"Season Player High: {info['player']} ({info['owner']}) {info['points']} pts")

            # Largest Blowout
            if val == "Largest Blowout / Margin of Victory":
                winner_cell = ws.cell(row=c.row, column=3)
                if not winner_cell.value and blowout.get('owner'):
                    winner_cell.value = blowout['owner']
                    ws.cell(row=c.row, column=5).value = f"+{blowout['margin']} pts (Week {blowout['week']})"
                    filled.append(f"Blowout: {blowout['owner']} +{blowout['margin']} (Wk {blowout['week']})")

            # Largest Win Streak
            if val == "Largest Regular Season Win Streak":
                winner_cell = ws.cell(row=c.row, column=3)
                if not winner_cell.value and streak.get('owner'):
                    winner_cell.value = streak['owner']
                    ws.cell(row=c.row, column=5).value = f"{streak['streak']} consecutive wins"
                    filled.append(f"Win Streak: {streak['owner']} ({streak['streak']}W)")

    if filled:
        print(f"\n✅ {year} ({len(filled)} cells filled):")
        for f in filled[:6]:
            print(f"   {f}")
        if len(filled) > 6:
            print(f"   ... and {len(filled)-6} more")
    else:
        print(f"\n⚠️  {year}: no data to prefill")

wb.save('GFN_Fantasy_Finances_v4.xlsx')
print("\n✅ Saved GFN_Fantasy_Finances_v4.xlsx")
